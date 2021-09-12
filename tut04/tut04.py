import csv
import os
# import shutil

from openpyxl import Workbook, load_workbook

# CSV input data file
REGISTRATION_CSV_FILE = "regtable_old.csv"
# field names for the output xlsx files
FIELDS = ["rollno", "register_sem", "subno", "sub_type"]


def create_folder(folder_name):
    """
    Create folder if it does not exist already
    """
    # if os.path.exists(folder_name):
    #     shutil.rmtree(folder_name)
    # os.makedirs(folder_name)
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)


def write_data(folder, key, reader):
    """
    Read csv data and write xlsx data
    into folder by key (for file name)
    """
    # key_idx denotes column index in CSV data
    for row in reader:
        # value present in the key field
        key_val = row[key]
        # skip writing if it is empty
        if not key_val:
            continue
        # name of the file, example: 1901EE01.xlsx
        file_name = key_val + ".xlsx"
        # file path with folder
        file_path = os.path.join(folder, file_name)
        # if file does not exist, create it
        # and insert field names
        if not os.path.exists(file_path):
            wb = Workbook()
            sheet = wb.active
            # if shhet is None, raise error
            if not sheet:
                raise Exception("Error fetching active sheet")
            # append data
            sheet.append(FIELDS)
            # save file
            wb.save(file_path)
        # load existing file and append data
        wb = load_workbook(file_path)
        sheet = wb.active
        # if shhet is None, raise error
        if not sheet:
            raise Exception("Error fetching active sheet")
        # append data
        sheet.append([row[field] for field in FIELDS])
        # save file
        wb.save(file_path)


def output_by_subject():
    """
    Key:    subject     (subno)
    Output: xlsx files  (excel)
    """
    create_folder("output_by_subject")
    reg_csv_file = open(REGISTRATION_CSV_FILE, "r")
    reader = csv.DictReader(reg_csv_file)
    write_data("output_by_subject", "subno", reader)
    reg_csv_file.close()


def output_individual_roll():
    """
    Key:    roll number (rollno)
    Output: xlsx files  (excel)
    """
    create_folder("output_individual_roll")
    reg_csv_file = open(REGISTRATION_CSV_FILE, "r")
    reader = csv.DictReader(reg_csv_file)
    write_data("output_individual_roll", "rollno", reader)
    reg_csv_file.close()


output_individual_roll()
output_by_subject()
