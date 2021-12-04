import os
import shutil

# file handling
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side

# web server
from flask import Flask, request, render_template, send_from_directory

# email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


OUTPUT_FOLDER = "marksheet"
app = Flask(__name__)
# RESPONSE_FILE = os.path.join("sample_input", "responses.csv")
RESPONSE_FILE = os.path.join(app.root_path, "uploads", "responses.csv")
# MASTER_ROLL_FILE = os.path.join("sample_input", "master_roll.csv")
MASTER_ROLL_FILE = os.path.join(app.root_path, "uploads", "master_roll.csv")
LOGO_FILE = os.path.join("static", "iitp_logo.png")
ANSWER_KEY = []
STUDENTS = {}
MARKING_SCHEME = [5.0, -1.0, 0]
SENDER_EMAIL = "cs384_2021@gmail.com"
SENDER_PASSWORD = "cs384_2021"


def load_roll_nums():
    global STUDENTS
    roll_file = open(MASTER_ROLL_FILE, "r")
    reader = csv.DictReader(roll_file)
    for row in reader:
        STUDENTS[row["roll"]] = {
            "name": row["name"],
            "timestamp": "",
            "score": 0,
            "google_score": 0,
            "email": "",
            "webmail": "",
            "phone": "",
            "answers": [],
            "stats": [0, 0, 0],
        }
    roll_file.close()


def generate_roll_num_wise_marksheet():
    global STUDENTS
    for roll in STUDENTS:
        wb = Workbook()
        ws = wb.active
        if not ws:
            raise Exception("Error getting active sheet")
        ws.title = "quiz"
        img = Image(LOGO_FILE)
        img.height = 80
        img.width = 650
        ws.add_image(img, "A1")
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 20

        title_font = Font(
            size=18, name="Century", bold=True, underline="single"
        )
        bold_font = Font(size=12, name="Century", bold=True)
        normal_font = Font(size=12, name="Century")
        correct_ans_font = Font(size=12, name="Century", color="008000")
        wrong_ans_font = Font(size=12, name="Century", color="FF0000")
        actual_ans_font = Font(size=12, name="Century", color="0000FF")
        side_border = Side(border_style="thin", color="000000")
        full_border = Border(
            top=side_border,
            right=side_border,
            bottom=side_border,
            left=side_border,
        )
        center_align = Alignment(horizontal="center")

        ws.merge_cells("A5:E5")

        ws["A5"] = "Mark Sheet"
        ws["A5"].font = title_font
        ws["A5"].alignment = center_align

        ws["A6"] = "Name:"
        ws["A6"].font = normal_font
        ws["A6"].alignment = Alignment(horizontal="right")
        ws["B6"] = STUDENTS[roll]["name"]
        ws["B6"].font = bold_font
        ws["D6"] = "Exam:"
        ws["D6"].alignment = Alignment(horizontal="right")
        ws["D6"].font = normal_font
        ws["E6"] = "quiz"
        ws["E6"].font = bold_font

        ws["A7"] = "Roll Number:"
        ws["A7"].font = normal_font
        ws["A7"].alignment = Alignment(horizontal="right")
        ws["B7"] = roll
        ws["B7"].font = bold_font

        ws["A9"] = ""
        ws["A9"].border = full_border
        ws["B9"] = "Right"
        ws["B9"].font = bold_font
        ws["B9"].border = full_border
        ws["B9"].alignment = center_align
        ws["C9"] = "Wrong"
        ws["C9"].font = bold_font
        ws["C9"].border = full_border
        ws["C9"].alignment = center_align
        ws["D9"] = "Not Attempted"
        ws["D9"].font = bold_font
        ws["D9"].border = full_border
        ws["D9"].alignment = center_align
        ws["E9"] = "Max"
        ws["E9"].font = bold_font
        ws["E9"].border = full_border
        ws["E9"].alignment = center_align

        ws["A10"] = "No."
        ws["A10"].font = bold_font
        ws["A10"].border = full_border
        ws["A10"].alignment = center_align
        ws["B10"] = STUDENTS[roll]["stats"][0]
        ws["B10"].font = correct_ans_font
        ws["B10"].border = full_border
        ws["B10"].alignment = center_align
        ws["C10"] = STUDENTS[roll]["stats"][1]
        ws["C10"].font = wrong_ans_font
        ws["C10"].border = full_border
        ws["C10"].alignment = center_align
        ws["D10"] = STUDENTS[roll]["stats"][2]
        ws["D10"].font = normal_font
        ws["D10"].border = full_border
        ws["D10"].alignment = center_align
        ws["E10"] = len(ANSWER_KEY)
        ws["E10"].font = normal_font
        ws["E10"].border = full_border
        ws["E10"].alignment = center_align

        ws["A11"] = "Marking"
        ws["A11"].font = bold_font
        ws["A11"].border = full_border
        ws["A11"].alignment = center_align
        ws["B11"] = MARKING_SCHEME[0]
        ws["B11"].font = correct_ans_font
        ws["B11"].border = full_border
        ws["B11"].alignment = center_align
        ws["C11"] = MARKING_SCHEME[1]
        ws["C11"].font = wrong_ans_font
        ws["C11"].border = full_border
        ws["C11"].alignment = center_align
        ws["D11"] = MARKING_SCHEME[2]
        ws["D11"].font = normal_font
        ws["D11"].border = full_border
        ws["D11"].alignment = center_align
        ws["E11"] = ""
        ws["E11"].border = full_border

        ws["A12"] = "Total"
        ws["A12"].font = bold_font
        ws["A12"].border = full_border
        ws["A12"].alignment = center_align
        ws["B12"] = STUDENTS[roll]["stats"][0] * MARKING_SCHEME[0]
        ws["B12"].font = correct_ans_font
        ws["B12"].border = full_border
        ws["B12"].alignment = center_align
        ws["C12"] = STUDENTS[roll]["stats"][1] * MARKING_SCHEME[1]
        ws["C12"].font = wrong_ans_font
        ws["C12"].border = full_border
        ws["C12"].alignment = center_align
        ws["D12"] = STUDENTS[roll]["stats"][2] * MARKING_SCHEME[2]
        ws["D12"].font = normal_font
        ws["D12"].border = full_border
        ws["D12"].alignment = center_align
        ws["E12"] = (
            str(STUDENTS[roll]["score"])
            + "/"
            + str(len(ANSWER_KEY) * MARKING_SCHEME[0])
        )
        ws["E12"].font = actual_ans_font
        ws["E12"].border = full_border
        ws["E12"].alignment = center_align

        two_columns = 1 if len(ANSWER_KEY) > 25 else 0
        ws["A15"] = "Student Ans"
        ws["A15"].font = bold_font
        ws["A15"].border = full_border
        ws["A15"].alignment = center_align
        ws["B15"] = "Correct Ans"
        ws["B15"].font = bold_font
        ws["B15"].border = full_border
        ws["B15"].alignment = center_align
        if two_columns:
            ws["D15"] = "Student Ans"
            ws["D15"].font = bold_font
            ws["D15"].border = full_border
            ws["D15"].alignment = center_align
            ws["E15"] = "Correct Ans"
            ws["E15"].font = bold_font
            ws["E15"].border = full_border
            ws["E15"].alignment = center_align

        for i in range(max(25, len(ANSWER_KEY))):
            if not STUDENTS[roll]["email"]:
                break
            ws["A" + str(16 + i)] = STUDENTS[roll]["answers"][i]
            if STUDENTS[roll]["answers"][i] == "":
                pass
            elif STUDENTS[roll]["answers"][i] == ANSWER_KEY[i]:
                ws["A" + str(16 + i)].font = correct_ans_font
            else:
                ws["A" + str(16 + i)].font = wrong_ans_font
            ws["A" + str(16 + i)].border = full_border
            ws["A" + str(16 + i)].alignment = center_align

            ws["B" + str(16 + i)] = ANSWER_KEY[i]
            ws["B" + str(16 + i)].font = actual_ans_font
            ws["B" + str(16 + i)].border = full_border
            ws["B" + str(16 + i)].alignment = center_align

            if two_columns and 25 + i < len(ANSWER_KEY):
                ws["D" + str(16 + i)] = STUDENTS[roll]["answers"][25 + i]
                if STUDENTS[roll]["answers"][25 + i] == "":
                    pass
                elif STUDENTS[roll]["answers"][25 + i] == ANSWER_KEY[25 + i]:
                    ws["D" + str(16 + i)].font = correct_ans_font
                else:
                    ws["D" + str(16 + i)].font = wrong_ans_font
                ws["D" + str(16 + i)].border = full_border
                ws["D" + str(16 + i)].alignment = center_align

                ws["E" + str(16 + i)] = ANSWER_KEY[25 + i]
                ws["E" + str(16 + i)].font = actual_ans_font
                ws["E" + str(16 + i)].border = full_border
                ws["E" + str(16 + i)].alignment = center_align

        wb.save(os.path.join(OUTPUT_FOLDER, roll + ".xlsx"))


def calculate_scores():
    global STUDENTS
    global MARKING_SCHEME
    resp_file = open(RESPONSE_FILE, "r")
    reader = csv.DictReader(
        resp_file,
        fieldnames=[
            "timestamp",
            "email",
            "score",
            "name",
            "webmail",
            "phone",
            "roll",
        ],
        restkey="answers",
    )
    for row in reader:
        if row["roll"] == "Roll Number":
            continue
        STUDENTS[row["roll"]]["answers"] = row["answers"]
        STUDENTS[row["roll"]]["email"] = row["email"]
        STUDENTS[row["roll"]]["webmail"] = row["webmail"]
        STUDENTS[row["roll"]]["timestamp"] = row["timestamp"]
        STUDENTS[row["roll"]]["google_score"] = row["score"]
        STUDENTS[row["roll"]]["phone"] = row["phone"]
        for idx, ans in enumerate(row["answers"]):
            if ans == "":
                STUDENTS[row["roll"]]["stats"][2] += 1
                STUDENTS[row["roll"]]["score"] += MARKING_SCHEME[2]
            elif ANSWER_KEY[idx] == ans:
                STUDENTS[row["roll"]]["stats"][0] += 1
                STUDENTS[row["roll"]]["score"] += MARKING_SCHEME[0]
            else:
                STUDENTS[row["roll"]]["stats"][1] += 1
                STUDENTS[row["roll"]]["score"] += MARKING_SCHEME[1]
    resp_file.close()


def clear():
    global STUDENTS
    global ANSWER_KEY
    STUDENTS = {}
    ANSWER_KEY = []
    if os.path.exists(OUTPUT_FOLDER):
        shutil.rmtree(OUTPUT_FOLDER)
    os.makedirs(OUTPUT_FOLDER)


def get_answer_key():
    global ANSWER_KEY
    resp_file = open(RESPONSE_FILE, "r")
    reader = csv.DictReader(
        resp_file,
        fieldnames=[
            "timestamp",
            "email",
            "score",
            "name",
            "webmail",
            "phone",
            "roll",
        ],
        restkey="answers",
    )
    flag = 0
    for row in reader:
        if row["roll"] == "ANSWER":
            for ans in row["answers"]:
                ANSWER_KEY.append(ans)
            flag += 1
        elif row["timestamp"] == "Timestamp":
            flag += 1
        if flag == 2:
            break
    if flag == 1:
        raise Exception(
            "No roll number with ANSWER is present, Cannot Process!"
        )
    resp_file.close()


@app.route("/download/<path:filename>")
def download_file(filename):
    return send_from_directory(os.path.join(OUTPUT_FOLDER, filename), filename)


@app.route("/sendemail")
def send_email():
    try:
        if not os.path.exists(OUTPUT_FOLDER) or len(os.listdir(OUTPUT_FOLDER)) == 0:
            return render_template(
                "index.html", error="Generate roll-num-wise marksheets first"
            )
        try:
            session = smtplib.SMTP("smtp.gmail.com", 587)
            session.starttls()
            session.login(SENDER_EMAIL, SENDER_PASSWORD)
        except:
            return render_template(
                "index.html", error="Error creating SMTP session"
            )
        for roll in STUDENTS:
            content = f'Hello {STUDENTS[roll]["name"]},\n\tPlease find your quiz marks attached to this email.'
            message = MIMEMultipart()
            message["From"] = SENDER_EMAIL
            message["To"] = (
                STUDENTS[roll]["email"] + "," + STUDENTS[roll]["webmail"]
            )
            message["Subject"] = "Quiz marks"
            message.attach(MIMEText(content, "plain"))
            part = MIMEBase("application", "octet-stream")
            with open(os.path.join(OUTPUT_FOLDER, roll + ".xlsx"), "rb") as f:
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition", f'attachment; filename="{roll}.xlsx"'
            )
            message.attach(part)
            session.sendmail(
                SENDER_EMAIL, STUDENTS[roll]["email"], message.as_string()
            )
        session.quit()
        return render_template("index.html", message="Emails sent")
    except Exception as e:
        return render_template("index.html", error=e)


@app.route("/concisescore")
def generate_concise_marksheet():
    try:
        if not os.path.exists(OUTPUT_FOLDER) or len(os.listdir(OUTPUT_FOLDER)) == 0:
            return render_template(
                "index.html", error="Generate roll-num-wise marksheets first"
            )
        concise_score_file = open(
            os.path.join(OUTPUT_FOLDER, "concise_marksheet.csv"), "w"
        )
        marksheet = csv.writer(concise_score_file)
        fields = [
            "Timestamp",
            "Email address",
            "Google_Score",
            "Name",
            "IITP webmail",
            "Phone (10 digit only)",
            "Score_After_Negative",
            "Roll Number",
        ]
        for _ in range(len(ANSWER_KEY)):
            fields.append("")
        fields.append("statusAns")
        marksheet.writerow(fields)
        for roll in STUDENTS:
            if not STUDENTS[roll]["email"]:
                row = ["ABSENT"]*(len(ANSWER_KEY) + 9)
                row[3] = STUDENTS[roll]["name"]
                row[7] = roll
                row[-1] = "[0, 0, " + str(len(ANSWER_KEY)) + "]"
                marksheet.writerow(row)
                continue
            row = [
                STUDENTS[roll]["timestamp"],
                STUDENTS[roll]["email"],
                STUDENTS[roll]["google_score"],
                STUDENTS[roll]["name"],
                STUDENTS[roll]["webmail"],
                STUDENTS[roll]["phone"],
                str(STUDENTS[roll]["score"])
                + " / "
                + str(len(ANSWER_KEY) * MARKING_SCHEME[0]),
                roll,
            ]
            for ans in STUDENTS[roll]["answers"]:
                row.append(ans)
            row.append(
                "["
                + str(STUDENTS[roll]["stats"][0])
                + ", "
                + str(STUDENTS[roll]["stats"][1])
                + ", "
                + str(STUDENTS[roll]["stats"][2])
                + "]"
            )
            marksheet.writerow(row)
        concise_score_file.close()
        return render_template("index.html", concise_generated=True)
    except Exception as e:
        return render_template("index.html", error=e)


@app.route("/", methods=["GET", "POST"])
def home_page():
    try:
        if request.method == "GET":
            return render_template("index.html")
        if not request.files["roll"]:
            return render_template("index.html", error="Roll file missing")
        if not request.files["response"]:
            return render_template("index.html", error="Response file missing")
        if not os.path.exists(os.path.join(app.root_path, "uploads")):
            os.makedirs(os.path.join(app.root_path, "uploads"))
        request.files["roll"].save(MASTER_ROLL_FILE)
        request.files["response"].save(RESPONSE_FILE)
        try:
            if request.form["pmarks"]:
                MARKING_SCHEME[0] = float(request.form["pmarks"])
            if request.form["nmarks"]:
                MARKING_SCHEME[1] = float(request.form["nmarks"])
        except:
            return render_template("index.html", error="Invalid marks entered")
        try:
            clear()
        except:
            print("Error clearing variables and old files")
        try:
            get_answer_key()
        except Exception as e:
            return render_template("index.html", error=e)
        try:
            load_roll_nums()
        except:
            return render_template(
                "index.html",
                error="Error reading master_roll csv and initializing STUDENTS dict",
            )
        try:
            calculate_scores()
        except:
            return render_template(
                "index.html", error="Error calculating scores"
            )
        generate_roll_num_wise_marksheet()
        return render_template("index.html")
    except Exception as e:
        return render_template("index.html", error=e)


if __name__ == "__main__":
    app.run()
